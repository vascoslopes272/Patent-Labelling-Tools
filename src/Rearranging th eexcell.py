import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def generate_super_compact_excel(input_file, output_xlsx):
    # 1. Load your dataset
    if input_file.endswith('.csv'):
        df = pd.read_csv(input_file)
    else:
        df = pd.read_excel(input_file)

    # 2. Create the workbook structure
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    # 3. Add stylish headers
    headers = list(df.columns)
    ws.append(headers)
    
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    header_font = Font(bold=True)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    # 4. Populate data and aggressively truncate unneeded text
    for _, row in df.iterrows():
        row_vals = []
        for col_name in df.columns:
            val = row[col_name]
            
            if pd.isna(val):
                row_vals.append(None)
                continue
                
            val_str = str(val)
            
            # Reduce Column E (Value) for fields you will never read
            if col_name == 'Value':
                field_val = str(row['Field'])
                if field_val in ['abstract', 'description_of_drawings']:
                    if len(val_str) > 25:
                        val_str = val_str[:25] + "..."
                        
            # Reduce Column G (Image_Path) to keep rows short for easier scrolling
            if col_name == 'Image_Path':
                if len(val_str) > 15:
                    val_str = val_str[:15] + "..."
            
            row_vals.append(val_str)
            
        ws.append(row_vals)

    # 5. Merge consecutive rows with identical values
    for col in range(1, ws.max_column + 1):
        start_row = 2
        for row in range(3, ws.max_row + 1):
            val_prev = ws.cell(row=row-1, column=col).value
            val_curr = ws.cell(row=row, column=col).value
            
            if val_curr != val_prev or val_prev is None:
                if (row - 1) > start_row and val_prev is not None:
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=row-1, end_column=col)
                    ws.cell(row=start_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
                start_row = row
                
        if ws.max_row > start_row and ws.cell(row=start_row, column=col).value is not None:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=ws.max_row, end_column=col)
            ws.cell(row=start_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)

    # 6. Set deliberate column widths (A, B, C, E, D are clear; F and G are small)
    column_widths = {
        'A': 18,  # Patent_ID
        'B': 10,  # Section
        'C': 30,  # Sub_Dimension
        'D': 22,  # Field
        'E': 50,  # Value (Width reduced slightly since text inside is much shorter now)
        'F': 10,  # Source
        'G': 12,  # Image_Path (Kept compact to hide long paths)
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 7. Final text-wrapping and alignment configurations
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            current_alignment = ws.cell(row=row, column=col).alignment
            ws.cell(row=row, column=col).alignment = Alignment(
                vertical='center' if current_alignment and current_alignment.vertical == 'center' else 'top',
                wrap_text=True
            )

    # 8. Save the file
    wb.save(output_xlsx)
    print(f"Successfully generated super compact file: {output_xlsx}")

# Run the script to produce your final optimized sheet
generate_super_compact_excel('reviewed_patents_Batch_05.xlsx - Review.csv', 'reviewed_patents_super_compact.xlsx')