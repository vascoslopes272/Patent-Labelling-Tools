"""
inject_pdf_links.py — add a per-patent `pdf_link` row to an ml_predict_labels
export, taken from the PatSeer source Excel's 'PDF Link' column.

WHY: the taxonomy review wizard opens a side-by-side source window on every
Next/prev. We want that window to show the PatSeer PDF (the actual document),
while the "Open in Espacenet (EPO)" button stays as the always-works fallback.
The PatSeer PDF URL lives ONLY in the source PatSeer Excel, stored as an Excel
*hyperlink* on the 'PDF Link' cell (the cell text is literally "PDF Link"; the
real URL is the hyperlink target). The ML export doesn't carry it, so we copy
it in here as one extra row per patent: Section=T1, Field="pdf_link".

The wizard reads it back with metaVal('pdf_link') (same lookup it uses for
title/abstract/etc.), so no schema change is needed — it's just one more row
in the existing flat Patent_ID/Section/Field/Value layout.

SAFE: backs up the export before writing (timestamped, never overwrites an
existing backup); idempotent (removes any prior pdf_link rows before adding,
so re-running doesn't duplicate); never touches the PatSeer source file.

Usage:
    python3 scripts/inject_pdf_links.py <ml_predict_labels_xlsx> [patseer_xlsx]

If patseer_xlsx is omitted, it's read from config.yaml (paths.patseer_excel).
"""

import shutil
import sys
import warnings
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

# Cosmetic: our new metadata rows have all-NA Confidence/Image_Path/Needs_Review
# columns, which trips a pandas concat deprecation warning. The behaviour is
# exactly what we want (keep df's columns), so silence just this one.
warnings.filterwarnings(
    "ignore",
    message="The behavior of DataFrame concatenation with empty or all-NA entries",
    category=FutureWarning,
)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def load_patseer_pdf_links(patseer_path: Path) -> dict[str, str]:
    """Return {Record Number: pdf_url} by reading the 'PDF Link' cell hyperlinks
    from the PatSeer Excel. pandas can't see hyperlinks, so use openpyxl."""
    wb = openpyxl.load_workbook(patseer_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if "PDF Link" not in header or "Record Number" not in header:
        raise SystemExit(
            f"PatSeer Excel missing required columns. Have: {header[:8]}..."
        )
    pdf_col = header.index("PDF Link") + 1
    rec_col = header.index("Record Number") + 1

    links: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        rec = ws.cell(row=r, column=rec_col).value
        cell = ws.cell(row=r, column=pdf_col)
        url = cell.hyperlink.target if cell.hyperlink else None
        if rec and url:
            links[str(rec).strip()] = url
    wb.close()
    return links


def inject(export_path: Path, links: dict[str, str]) -> None:
    df = pd.read_excel(export_path)

    # Drop any pre-existing pdf_link rows so re-running is idempotent.
    before = len(df)
    df = df[df["Field"] != "pdf_link"].copy()
    removed = before - len(df)

    # Build one pdf_link row per patent, mirroring the existing T1 metadata
    # row shape (Section=T1) so the wizard's metaVal() finds it.
    new_rows = []
    matched = 0
    pids = list(dict.fromkeys(df["Patent_ID"].dropna().tolist()))  # preserve order, unique
    for pid in pids:
        url = links.get(str(pid).strip())
        if not url:
            continue
        matched += 1
        new_rows.append({
            "Patent_ID": pid, "Section": "T1", "Sub_Dimension": "PDF Link",
            "Field": "pdf_link", "Definition": "PatSeer source PDF", "Options": "",
            "Value": url, "Confidence": None, "Source": "patseer",
            "Image_Path": None, "Needs_Review": None,
        })

    # Align the new rows to the existing columns before concat so pandas doesn't
    # warn about all-NA columns (Confidence/Image_Path/Needs_Review are None here)
    # and the result dtypes stay anchored to df. No-op when nothing matched.
    if new_rows:
        new_df = pd.DataFrame(new_rows).reindex(columns=df.columns)
        out = pd.concat([df, new_df], ignore_index=True)
    else:
        out = df

    # Backup before writing (timestamped, never clobber an existing backup).
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = export_path.with_name(export_path.stem + f".PRE_PDFLINK_{ts}.xlsx")
    shutil.copy2(export_path, backup)

    out.to_excel(export_path, sheet_name="Review", index=False)

    print(f"  removed {removed} stale pdf_link row(s)")
    print(f"  patents in export      : {len(pids)}")
    print(f"  pdf_link rows added    : {matched}")
    print(f"  patents with NO PatSeer link: {len(pids) - matched}")
    print(f"  backup                 : {backup}")
    print(f"  written                : {export_path}")


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    export_path = Path(sys.argv[1])
    if not export_path.exists():
        raise SystemExit(f"Export not found: {export_path}")

    if len(sys.argv) >= 3:
        patseer_path = Path(sys.argv[2])
    else:
        from src.config_loader import load_config
        patseer_path = Path(load_config()["paths"]["patseer_excel"])
    if not patseer_path.exists():
        raise SystemExit(f"PatSeer Excel not found: {patseer_path}")

    print(f"PatSeer source: {patseer_path}")
    print(f"Export        : {export_path}")
    links = load_patseer_pdf_links(patseer_path)
    print(f"  PatSeer PDF links found: {len(links)}")
    inject(export_path, links)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
